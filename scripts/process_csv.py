import csv
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def main():
    """主函数，处理命令行参数并执行处理逻辑"""
    # GitHub友好型参数处理
    if len(sys.argv) < 2:
        print("使用方法: python script.py [输入CSV路径] [输出XLSX路径]")
        print("示例: python script.py data/input.csv results/output.xlsx")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else "results/output.xlsx"
    
    # 确保输出目录存在
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    # 处理CSV并生成Excel
    process_csv(input_file, output_file)

def process_csv(input_file, output_file):
    """
    处理CSV文件，计算积分和净胜值，并生成带排名的Excel文件
    
    参数:
        input_file (str): 输入CSV文件路径
        output_file (str): 输出Excel文件路径
    """
    # 定义列索引（0-based）
    C_INDEX = 2    # C列
    D_INDEX = 3    # D列
    M_INDEX = 12   # M列
    N_INDEX = 13   # N列
    O_INDEX = 14   # O列
    P_INDEX = 15   # P列
    Q_INDEX = 16   # Q列
    R_INDEX = 17   # R列
    S_INDEX = 18   # S列
    
    try:
        # 初始化队伍统计
        team_stats = {}
        
        # 创建Excel工作簿
        wb = Workbook()
        ws_detail = wb.active
        ws_detail.title = "详细数据"
        ws_ranking = wb.create_sheet(title="队伍排名")
        
        # 添加标题行
        detail_headers = ["队伍A", "队伍B", "积分A", "净胜A", "积分B", "净胜B"]
        ws_detail.append(detail_headers)
        
        ranking_headers = ["排名", "队伍名称", "总积分", "总净胜分", "参赛场次"]
        ws_ranking.append(ranking_headers)
        
        # 读取并处理CSV文件
        with open(input_file, 'r', newline='', encoding='utf-8') as infile:
            reader = csv.reader(infile)
            
            # 跳过标题行（如果有）
            try:
                headers = next(reader)
                # 如果CSV有标题，更新详细数据表的列名
                if len(headers) > max(C_INDEX, D_INDEX):
                    detail_headers[0] = headers[C_INDEX]
                    detail_headers[1] = headers[D_INDEX]
            except StopIteration:
                pass  # 空文件
            
            # 处理每一行数据
            for row_idx, row in enumerate(reader, start=1):
                # 跳过空行
                if not row:
                    continue
                
                # 提取队伍数据
                c_val = row[C_INDEX] if len(row) > C_INDEX else f"队伍{row_idx}A"
                d_val = row[D_INDEX] if len(row) > D_INDEX else f"队伍{row_idx}B"
                
                # 初始化变量
                m = n = o = p = r = s = 0.0
                q = ""
                
                # 提取其他列数据
                try:
                    if len(row) > M_INDEX and row[M_INDEX].strip():
                        m = float(row[M_INDEX])
                    if len(row) > N_INDEX and row[N_INDEX].strip():
                        n = float(row[N_INDEX])
                    if len(row) > O_INDEX and row[O_INDEX].strip():
                        o = float(row[O_INDEX])
                    if len(row) > P_INDEX and row[P_INDEX].strip():
                        p = float(row[P_INDEX])
                    if len(row) > Q_INDEX:
                        q = row[Q_INDEX].strip()
                    if len(row) > R_INDEX and row[R_INDEX].strip():
                        r = float(row[R_INDEX])
                    if len(row) > S_INDEX and row[S_INDEX].strip():
                        s = float(row[S_INDEX])
                except (ValueError, TypeError) as e:
                    print(f"行 {row_idx}: 数据转换错误 - {str(e)}")
                    continue
                
                # 计算积分和净胜值
                score_a, net_a, score_b, net_b = calculate_scores(m, n, o, p, q, r, s)
                
                # 写入详细数据
                ws_detail.append([c_val, d_val, score_a, net_a, score_b, net_b])
                
                # 更新队伍统计
                update_team_stats(team_stats, c_val, score_a, net_a)
                update_team_stats(team_stats, d_val, score_b, net_b)
        
        # 生成排名表
        generate_ranking(ws_ranking, team_stats)
        
        # 美化Excel格式
        style_excel(wb)
        
        # 保存Excel文件
        wb.save(output_file)
        print(f"处理完成！结果已保存至 {output_file}")
        print(f"处理行数: {row_idx}")
        print(f"队伍数量: {len(team_stats)}")
    
    except Exception as e:
        print(f"处理过程中出错: {str(e)}")
        import traceback
        traceback.print_exc()

def calculate_scores(m, n, o, p, q, r, s):
    """根据规则计算积分和净胜值"""
    score_a = 0
    score_b = 0
    net_a = 0
    net_b = 0
    
    # 规则1: 比较M和N
    if m != n:
        if m > n:
            score_a = 1
            net_a = m - n
            net_b = -(m - n)
        else:
            score_b = 1
            net_b = n - m
            net_a = -(n - m)
    
    # 规则2: M等于N时比较O和P
    elif o != p:
        if o > p:
            score_a = 1
        else:
            score_b = 1
        # 净胜值保持为0
    
    # 规则3: O等于P时处理Q,R,S
    else:
        if q == "四抓":
            # 四抓规则: 小的一方得1分
            if r < s:
                score_a = 1
            elif r > s:
                score_b = 1
        else:
            # 非四抓规则: 大的一方得1分
            if r > s:
                score_a = 1
            elif r < s:
                score_b = 1
        # 净胜值保持为0
    
    return score_a, net_a, score_b, net_b

def update_team_stats(team_stats, team_name, score, net):
    """更新队伍统计数据"""
    if not team_name:
        return
    
    if team_name not in team_stats:
        team_stats[team_name] = {"score": 0, "net": 0, "games": 0}
    
    team_stats[team_name]["score"] += score
    team_stats[team_name]["net"] += net
    team_stats[team_name]["games"] += 1

def generate_ranking(ws_ranking, team_stats):
    """生成排名表"""
    # 准备排名数据
    ranking_data = []
    for team, stats in team_stats.items():
        ranking_data.append((team, stats["score"], stats["net"], stats["games"]))
    
    # 按积分和净胜分排序
    ranking_data.sort(key=lambda x: (-x[1], -x[2]))
    
    # 添加排名并分配名次
    current_rank = 1
    prev_score = None
    prev_net = None
    skip_rank = 0  # 用于处理并列排名后的跳过
    
    for i, (team, score, net, games) in enumerate(ranking_data):
        # 处理并列排名
        if prev_score == score and prev_net == net:
            # 并列排名，使用相同名次
            rank_display = current_rank
            skip_rank += 1
        else:
            # 新排名 = 当前位置 + 1 - 跳过的排名数
            current_rank = i + 1 - skip_rank
            rank_display = current_rank
        
        ws_ranking.append([rank_display, team, score, net, games])
        
        prev_score = score
        prev_net = net

def style_excel(wb):
    """简化Excel格式美化"""
    # 基础样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    for sheet in wb:
        # 跳过空工作表
        if sheet.max_row == 0:
            continue
            
        # 设置列宽
        for col in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            
            for cell in col:
                try:
                    value_length = len(str(cell.value))
                    if value_length > max_length:
                        max_length = value_length
                except:
                    pass
            
            # 设置列宽，留出一些余量
            sheet.column_dimensions[column_letter].width = max_length + 2
        
        # 设置标题行样式
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 设置数据行样式
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                
                # 数值列右对齐
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")
        
        # 冻结标题行
        if sheet.max_row > 1:
            sheet.freeze_panes = "A2"

if __name__ == "__main__":
    main()
